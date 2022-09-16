import os
import re
from openpyxl import Workbook

provinceList = [
    "北京", "天津", "上海", "重庆", "内蒙古", "广西", "西藏", "宁夏", "新疆", "河北", "山西", "辽宁", "吉林",
    "黑龙江", "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "河南", "海南",
    "四川", "贵州", "云南", "陕西", "甘肃", "青海"]

# 定义一个全局变量字典，存放每一天的新增确诊和新增无症状
alldata = dict()


def process_reg(province, d, str):
    #如果在province里找不到逗号，说明当日确诊均在某一个省
    #直接遍历provinceList里的省份，如果某个省份出现在province里
    #就把该省份作为key，当日新增确诊作为value存入字典
    if province.find('，') == -1:
        for name in provinceList:
            if name in province:
                d[name] = int(d[str])
                return
    else:
        #如果找到逗号，且第一个字是“在”或者“均”，则说明也是上面一种情况
        #处理方式同上
        if province.find("均") == 0 or province.find("在") == 0:
            for name in provinceList:
                if name in province:
                    d[name] = int(d[str])
                    return
        #如果不是，则说明有多个省份有新增确诊，遍历省份列表
        for name in provinceList:
            #编译正则表达式，匹配 <省份名><任意多个数字>例 这种模式的数据
            pattern = re.compile(r"%s[0-9]*例" % name)
            data = re.search(pattern, province)
            #如果找到该省份的数据
            if data:
                data = data.group()
                #获取数据后面的病例数并存入字典
                d[name] = int(data[data.find(name) + len(name):data.find('例')])


def process_new(filenames, basePath, str):
    for filename in filenames:
        # 截取文件名前10个字符，得到文件的日期
        date = filename[0:10]

        d = dict()
        # 如果此时为判断新增确诊，则以日期为key在全局变量里建一个字典
        if str == '新增确诊':
            alldata[date] = dict()
        # 打开对应文件
        file = open(basePath + "\\" + filename, 'r', encoding='utf-8')
        content = file.read()
        # 编译正则表达式，得到文本中有关新增确诊或新增无症状的信息
        reg = r'本土病例.*?）' if str == '新增确诊' else r'本土\d.*）(?=。\n当日解除|；当日解除|；当日无)'
        pattern = re.compile(reg)
        result = re.search(pattern, content)

        #
        if result:
            # 求出当日本土新增总病例
            result = result.group()
            start = 4 if str == '新增确诊' else 2
            end = "例（" if str == '新增确诊' else "例"
            d[str] = int(result[start:result.find(end)])
            # 求出各省份新增病例
            provinces = result
            # 把括号里的内容提取出来，便于数据处理
            provinces = provinces[provinces.index('（') + 1:provinces.index("）")]

            # 处理数据，把provinces里各省份与其新增确诊或新增无症状病例
            # 作为key，value存入alldata[date][str]
            process_reg(provinces, d, str)

        else:
            #找不到，说明当日新增确诊或新增无症状为零
            d[str] = 0
        alldata[date][str] = dict()
        alldata[date][str].update(d)


def pipe_into_excel(date):
    wb = Workbook()
    sheet1 = wb.create_sheet(index=1, title="新增确诊")
    sheet2 = wb.create_sheet(index=2, title="新增无症状")
    i = 1

    for key, value in alldata[date]['新增确诊'].items():
        sheet1.cell(i, 1, value=key)
        sheet1.cell(i, 2, value=value)
        i = i + 1
    i = 1
    for key, value in alldata[date]['新增无症状'].items():
        sheet2.cell(i, 1, value=key)
        sheet2.cell(i, 2, value=value)
        i = i + 1

    wb.save('D:\\excel\\' + date + '.xlsx')


if "__main__" == __name__:
    basePath = "D:\\疫情防控数据"
    # 读出所有数据
    filenames = os.listdir(basePath)
    # 收集新增确诊数据
    process_new(filenames, basePath, '新增确诊')
    # 收集新增无症状数据
    process_new(filenames, basePath, '新增无症状')
    print(alldata)
    # 将数据导入excel
    for key, value in alldata.items():
        pipe_into_excel(key)
